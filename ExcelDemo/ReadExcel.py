import openpyxl

wb=openpyxl.load_workbook("C:/Users/Nakul Sudhakaran/Desktop/Book1.xlsx")
sheets=wb.sheetnames
print(sheets)
print(wb.active.title)
sh1=wb['India']
data=sh1['A2'].value
data2=wb['India']['B2'].value
print(data)
print(data2)

sh2=wb['Kannur']
print(sh2.cell(2,1).value)

#print(wb.get_sheet_by_name('Names').cell(2,2).value)       Deprecated in new version of Python


