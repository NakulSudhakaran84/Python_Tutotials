import openpyxl

wb=openpyxl.load_workbook("Data.xlsx")
sheets=wb['Names']
row=sheets.max_row
column=sheets.max_column

for i in range(2,row+1):
    for j in range(1,column+1):
        print(sheets.cell(i,j).value)


sheets.cell(row=5,column=1,value='Java')
sheets.cell(row=5,column=2,value='Kerala')
sheets.cell(row=5,column=3,value=80)

wb.save("Data.xlsx")